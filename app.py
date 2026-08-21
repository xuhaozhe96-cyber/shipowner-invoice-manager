from __future__ import annotations

import io
import json
import mimetypes
import os
import re
import sqlite3
import sys
import threading
import uuid
import webbrowser
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
PAYMENT_PROOF_DIR = DATA_DIR / "payment_proofs"
DATABASE = DATA_DIR / "shipowner_invoices.db"
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
DEFAULT_COMPANY_NAME = "Cargo Move Ltd"
DEFAULT_BADGE = "GKM"
COSCO_RELEASES_EMAIL = "Releases@coscoshipping.co.uk"
PAYMENT_PROOF_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".gif"}

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.environ.get("SECRET_KEY", "local-development-key-change-me"),
    MAX_CONTENT_LENGTH=MAX_UPLOAD_BYTES,
)

EDITABLE_FIELDS = [
    "vessel_name",
    "voyage_no",
    "eta",
    "invoice_no",
    "invoice_date",
    "owner_name",
    "owner_email",
    "port_of_discharge",
    "container_no",
    "container_size",
    "bl_no",
    "charge_details",
    "amount",
    "currency",
    "status",
]

TEMPLATE_LEARN_FIELDS = [
    "vessel_name",
    "voyage_no",
    "eta",
    "invoice_no",
    "invoice_date",
    "owner_name",
    "owner_email",
    "port_of_discharge",
    "container_no",
    "container_size",
    "bl_no",
    "amount",
    "currency",
]

FIELD_LABELS = {
    "vessel_name": "船名",
    "voyage_no": "航次",
    "eta": "ETA",
    "invoice_no": "账单号",
    "invoice_date": "账单日期",
    "owner_name": "船东",
    "owner_email": "邮箱",
    "port_of_discharge": "卸货港",
    "container_no": "集装箱号",
    "container_size": "集装箱尺寸",
    "bl_no": "提单号",
    "amount": "金额",
    "currency": "币种",
}


def get_db() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    PAYMENT_PROOF_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DATABASE)
    connection.row_factory = sqlite3.Row
    return connection


@contextmanager
def db_session():
    """Commit successful work and always release the SQLite file handle."""
    connection = get_db()
    try:
        yield connection
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def init_db() -> None:
    with db_session() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                raw_text TEXT NOT NULL DEFAULT '',
                extraction_warning TEXT NOT NULL DEFAULT '',
                learning_note TEXT NOT NULL DEFAULT '',
                invoice_category TEXT NOT NULL DEFAULT 'freight',
                vessel_name TEXT NOT NULL DEFAULT '',
                voyage_no TEXT NOT NULL DEFAULT '',
                eta TEXT NOT NULL DEFAULT '',
                invoice_no TEXT NOT NULL DEFAULT '',
                invoice_date TEXT NOT NULL DEFAULT '',
                owner_name TEXT NOT NULL DEFAULT '',
                owner_email TEXT NOT NULL DEFAULT '',
                port_of_discharge TEXT NOT NULL DEFAULT '',
                container_no TEXT NOT NULL DEFAULT '',
                container_size TEXT NOT NULL DEFAULT '',
                bl_no TEXT NOT NULL DEFAULT '',
                charge_details TEXT NOT NULL DEFAULT '',
                amount TEXT NOT NULL DEFAULT '',
                currency TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '待校正',
                email_subject TEXT NOT NULL DEFAULT '',
                email_body TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        existing_columns = {row[1] for row in db.execute("PRAGMA table_info(invoices)").fetchall()}
        for column in (
            "port_of_discharge",
            "container_size",
            "charge_details",
            "invoice_date",
            "learning_note",
            "invoice_category",
        ):
            if column not in existing_columns:
                db.execute(f"ALTER TABLE invoices ADD COLUMN {column} TEXT NOT NULL DEFAULT ''")
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS email_drafts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vessel_name TEXT NOT NULL,
                owner_name TEXT NOT NULL DEFAULT '',
                owner_email TEXT NOT NULL DEFAULT '',
                subject TEXT NOT NULL DEFAULT '',
                body TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '草稿已保存',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(vessel_name, owner_name, owner_email)
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS payment_proofs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vessel_name TEXT NOT NULL,
                owner_name TEXT NOT NULL DEFAULT '',
                payment_date TEXT NOT NULL DEFAULT '',
                source_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS cosco_release_drafts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vessel_name TEXT NOT NULL,
                owner_name TEXT NOT NULL DEFAULT '',
                owner_email TEXT NOT NULL DEFAULT '',
                subject TEXT NOT NULL DEFAULT '',
                body TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(vessel_name, owner_name, owner_email)
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS vessel_archives (
                vessel_name TEXT PRIMARY KEY,
                archived_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS container_releases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vessel_name TEXT NOT NULL,
                owner_name TEXT NOT NULL DEFAULT '',
                container_no TEXT NOT NULL,
                released INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL,
                UNIQUE(vessel_name, owner_name, container_no)
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS invoice_learning_examples (
                invoice_id INTEGER PRIMARY KEY,
                source_filename TEXT NOT NULL DEFAULT '',
                raw_text TEXT NOT NULL,
                template_signature TEXT NOT NULL,
                fields_json TEXT NOT NULL,
                confirmed_fields TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS container_free_days (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vessel_name TEXT NOT NULL,
                owner_name TEXT NOT NULL DEFAULT '',
                container_no TEXT NOT NULL,
                last_free_day TEXT NOT NULL DEFAULT '',
                pickup_date TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                UNIQUE(vessel_name, owner_name, container_no)
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS free_day_extension_drafts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vessel_name TEXT NOT NULL,
                owner_name TEXT NOT NULL DEFAULT '',
                owner_email TEXT NOT NULL DEFAULT '',
                subject TEXT NOT NULL DEFAULT '',
                body TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(vessel_name, owner_name, owner_email)
            )
            """
        )
        seed_learning_examples(db)


def clean_value(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip(" :-\t")


def normalize_invoice_number(value: str) -> str:
    """Compare invoice numbers without case, spaces, slashes, or punctuation."""
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())


def find_duplicate_invoice(
    db: sqlite3.Connection, invoice_no: str, exclude_id: int | None = None
) -> sqlite3.Row | None:
    invoice_key = normalize_invoice_number(invoice_no)
    if not invoice_key:
        return None
    rows = db.execute(
        "SELECT * FROM invoices WHERE id <> ? ORDER BY id",
        (exclude_id or 0,),
    ).fetchall()
    return next(
        (
            row
            for row in rows
            if normalize_invoice_number(row["invoice_no"]) == invoice_key
        ),
        None,
    )


def first_match(text: str, patterns: list[str], flags: int = re.IGNORECASE) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return clean_value(match.group(1))
    return ""


def normalize_date(value: str) -> str:
    value = clean_value(value)
    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d, %Y",
        "%B %d, %Y",
    )
    for date_format in formats:
        try:
            return datetime.strptime(value, date_format).date().isoformat()
        except ValueError:
            pass
    return value


def display_decimal(value: str) -> str:
    try:
        number = Decimal(value.replace(",", ""))
    except (InvalidOperation, AttributeError):
        return value
    return f"{number:f}".rstrip("0").rstrip(".")


def extract_charge_details(text: str, currency: str) -> str:
    """Rebuild line-item descriptions from COSCO's column-ordered PDF text."""
    lines = [clean_value(line) for line in text.splitlines() if clean_value(line)]
    subtotal_index = next((i for i, line in enumerate(lines) if line.upper() == "SUB-TOTAL"), -1)
    if subtotal_index < 1:
        return ""

    names = []
    for line in reversed(lines[:subtotal_index]):
        if re.fullmatch(r"[-\d.,%]+", line) or "VATBASIS" in line.upper():
            break
        if len(line) <= 100:
            names.append(line)
    names.reverse()
    if not names or len(names) > 20:
        return ""

    count = len(names)
    amounts: list[str] = []
    upper_currency = (currency or "GBP").upper()
    for index in range(subtotal_index + 1, len(lines) - count * 3 + 1):
        if all(lines[index + offset].upper() == upper_currency for offset in range(count)):
            rate_start = index + count * 2
            candidates = lines[rate_start : rate_start + count]
            if len(candidates) == count and all(re.fullmatch(r"[\d,]+(?:\.\d+)?", value) for value in candidates):
                amounts = candidates
                break
    if len(amounts) != count:
        return "\n".join(names)

    symbol = {"GBP": "£", "USD": "$", "EUR": "€", "CNY": "¥", "RMB": "¥"}.get(upper_currency, f"{upper_currency} ")
    return "\n".join(f"{name} | {symbol}{display_decimal(amount)}" for name, amount in zip(names, amounts))


def extract_fields(text: str) -> dict[str, str]:
    """Extract common invoice fields. Every value remains editable by the user."""
    fields = {
        "vessel_name": first_match(
            text,
            [
                r"(?:vessel(?:\s*name)?|m[./]?v[.]?|船名)\s*[:#-]?\s*([^\n\r]+)",
                r"(?:vsl)\s*[:#-]?\s*([^\n\r]+)",
            ],
        ),
        "voyage_no": first_match(
            text,
            [r"(?:voyage(?:\s*(?:no\.?|number))?|voy\.?|航次)\s*[:#-]?\s*([A-Z0-9][A-Z0-9/.-]*)"],
        ),
        "eta": normalize_date(
            first_match(
                text,
                [
                    r"(?:ETA|预计到港(?:日期|时间)?)\s*[:#-]?\s*((?:\d{1,2}[./-]){2}\d{2,4})",
                    r"(?:ETA|预计到港(?:日期|时间)?)\s*[:#-]?\s*([A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})",
                ],
            )
        ),
        "invoice_no": first_match(
            text,
            [r"(?:invoice\s*(?:no\.?|number|#)?|账单号|发票号)\s*[:#-]?\s*([A-Z0-9][A-Z0-9/._-]*)"],
        ),
        "invoice_date": normalize_date(
            first_match(
                text,
                [
                    r"(?:invoice\s+date|issue\s+date|账单日期|发票日期)\s*[:#-]?\s*((?:\d{1,2}[./-]){2}\d{2,4})",
                    r"(?:invoice\s+date|issue\s+date|账单日期|发票日期)\s*[:#-]?\s*(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})",
                ],
            )
        ),
        "owner_name": first_match(
            text,
            [r"(?:shipowner|owner|carrier|船东|承运人)(?!\s+(?:security|charge))\s*[:#-]?\s*([^\n\r]+)"],
        ),
        "owner_email": first_match(
            text,
            [r"(?:e-?mail|邮箱)?\s*[:#-]?\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})"],
        ),
        "port_of_discharge": first_match(
            text,
            [r"(?:port\s+of\s+discharge|discharge\s+port|卸货港)\s*[:#-]?\s*([^\n\r]+)"],
        ),
        "container_no": first_match(
            text,
            [
                r"(?:container(?:\s*(?:no\.?|number|#))?|箱号)\s*[:#-]?\s*([A-Z]{4}\s*\d{7}(?:\s*[,;/]\s*[A-Z]{4}\s*\d{7})*)",
                r"\b([A-Z]{4}\s*\d{7})\b",
            ],
        ),
        "container_size": first_match(
            text,
            [r"\b[A-Z]{4}\s*\d{7}\s+(\d{2}(?:HQ|HC|GP|RF|OT|FR))\b"],
        ),
        "bl_no": first_match(
            text,
            [r"(?:B/?L|bill\s+of\s+lading|提单号)\s*(?:no\.?|number|#)?\s*[:#-]?\s*([A-Z0-9][A-Z0-9/._-]*)"],
        ),
        "currency": first_match(
            text,
            [r"(?:currency|币种)\s*[:#-]?\s*(USD|EUR|GBP|CNY|RMB|HKD|AUD|CAD)"],
        ),
        "amount": first_match(
            text,
            [
                r"(?:total\s*(?:amount)?|amount\s*due|grand\s*total|总金额|应付金额)\s*[:#-]?\s*(?:USD|EUR|GBP|CNY|RMB|HKD|AUD|CAD|[$€£¥])?\s*([\d,]+(?:\.\d{1,2})?)",
            ],
        ),
        "charge_details": "",
    }
    if not fields["currency"]:
        fields["currency"] = first_match(
            text,
            [r"(?:total\s*(?:amount)?|amount\s*due|grand\s*total)\s*[:#-]?\s*(USD|EUR|GBP|CNY|RMB|HKD|AUD|CAD)\b"],
        )
    fields["container_no"] = re.sub(r"\s+", "", fields["container_no"])

    # COSCO invoices extract in column order rather than visual reading order.
    # These reverse-label rules are intentionally scoped to that layout.
    if "COSCO SHIPPING Lines (UK) Limited" in text:
        fields["owner_name"] = "COSCO SHIPPING Lines"
        fields["owner_email"] = COSCO_RELEASES_EMAIL
        fields["invoice_no"] = first_match(text, [r"\b(\d{8,})\s*INVOICE\s+NO\."])
        fields["invoice_date"] = normalize_date(
            first_match(text, [r"\b(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})\s*ISSUE DATE"])
        )
        fields["vessel_name"] = first_match(
            text,
            [r"ARRIVED/DEPARTED\s*\n?\s*\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\s*([A-Z][A-Z ]+?\s+\d{2,4}[A-Z])(?:\s|$)"],
        )
        fields["voyage_no"] = ""
        fields["port_of_discharge"] = first_match(
            text,
            [r"PORT OF DISCHARGE[^\n]*\n([A-Za-z][A-Za-z .'-]+?)\s+\1(?:[A-Z]|\s|$)"],
        )
        fields["bl_no"] = first_match(text, [r"\n[A-Za-z][A-Za-z ]+\s+(\d{8,})(?=[A-Z][a-z])"])
        fields["currency"] = first_match(text, [r"\b(GBP|USD|EUR|CNY|RMB)\s*[\d,]+(?:\.\d{2})?\s*AMOUNT DUE"])
        fields["amount"] = first_match(text, [r"\b(?:GBP|USD|EUR|CNY|RMB)\s*([\d,]+(?:\.\d{2})?)\s*AMOUNT DUE"])
        fields["charge_details"] = extract_charge_details(text, fields["currency"])
    if "COSCO" in fields["owner_name"].upper():
        fields["owner_email"] = COSCO_RELEASES_EMAIL
    return fields


def extract_pdf_text(path: Path) -> tuple[str, str]:
    try:
        reader = PdfReader(path)
        pages = [(page.extract_text() or "").strip() for page in reader.pages]
        text = "\n\n".join(page for page in pages if page)
    except Exception as exc:  # malformed/encrypted PDFs vary by producer
        return "", f"PDF 无法读取：{exc}"
    if not text.strip():
        return "", "未提取到文字。此文件可能是扫描件，请手工填写字段。"
    return text, ""


def template_line_signature(line: str) -> str:
    """Remove changing invoice values while keeping the PDF layout wording."""
    value = clean_value(line).upper()
    value = re.sub(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", "<EMAIL>", value)
    value = re.sub(r"\b[A-Z]{4}\s*\d{7}\b", "<CONTAINER>", value)
    value = re.sub(
        r"\b\d{1,2}\s+(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\s+\d{4}(?=\b|[A-Z])",
        "<DATE>",
        value,
    )
    value = re.sub(r"\b\d{1,4}[./-]\d{1,2}[./-]\d{1,4}\b", "<DATE>", value)
    value = re.sub(r"\d[\d,.]*", "<N>", value)
    return re.sub(r"\s+", " ", value).strip()


def document_template_signature(text: str) -> str:
    lines = [template_line_signature(line) for line in text.splitlines() if clean_value(line)]
    return "\n".join(lines[:400])


def template_similarity(first: str, second: str) -> float:
    first_lines = first.splitlines()
    second_lines = second.splitlines()
    if not first_lines or not second_lines:
        return 0.0
    sequence_score = SequenceMatcher(None, first_lines, second_lines, autojunk=False).ratio()
    first_set, second_set = set(first_lines), set(second_lines)
    union = first_set | second_set
    set_score = len(first_set & second_set) / len(union) if union else 0.0
    return sequence_score * 0.7 + set_score * 0.3


def learning_field_candidates(field: str, line: str) -> list[str]:
    patterns: dict[str, list[str]] = {
        "vessel_name": [
            r"\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\s*([A-Z][A-Z .'&/-]{3,}\s+\d{2,4}[A-Z])\b",
            r"(?:VESSEL(?:\s*NAME)?|VSL)\s*[:#-]?\s*([A-Z][A-Z .'&/-]{3,}\s+\d{2,4}[A-Z])\b",
            r"^\s*([A-Z][A-Z .'&/-]{3,}\s+\d{2,4}[A-Z])\b",
        ],
        "voyage_no": [r"\b\d{2,4}[A-Z]\b"],
        "eta": [
            r"\b\d{4}-\d{1,2}-\d{1,2}\b",
            r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b",
            r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}(?=\b|[A-Z])",
        ],
        "invoice_date": [
            r"\b\d{4}-\d{1,2}-\d{1,2}\b",
            r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b",
            r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}(?=\b|[A-Z])",
        ],
        "invoice_no": [r"\d{6,}", r"\b[A-Z]{1,8}[-/]?\d[A-Z0-9/._-]{3,}\b"],
        "owner_email": [r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}"],
        "container_no": [r"\b[A-Z]{4}\s*\d{7}\b"],
        "container_size": [r"\b\d{2}(?:HQ|HC|GP|RF|OT|FR)\b"],
        "bl_no": [r"\d{7,}", r"\b[A-Z]{2,8}[-/]?\d[A-Z0-9/._-]{5,}\b"],
        "amount": [r"\d[\d,]*\.\d{1,4}"],
        "currency": [r"\b(?:GBP|USD|EUR|CNY|RMB|HKD|AUD|CAD)\b"],
    }
    candidates: list[str] = []
    for pattern in patterns.get(field, []):
        for match in re.finditer(pattern, line, re.IGNORECASE):
            candidate = clean_value(match.group(1) if match.lastindex else match.group(0))
            if candidate and candidate not in candidates:
                candidates.append(candidate)
    return candidates


def comparable_learning_value(field: str, value: str) -> str:
    value = clean_value(value)
    if field in {"eta", "invoice_date"}:
        return normalize_date(value)
    if field == "amount":
        try:
            return str(Decimal(value.replace(",", "")).normalize())
        except InvalidOperation:
            pass
    return re.sub(r"\s+", "", value).upper()


def locate_learning_value(field: str, value: str, text: str) -> tuple[int, int | None] | None:
    expected = comparable_learning_value(field, value)
    if not expected:
        return None
    lines = [clean_value(line) for line in text.splitlines() if clean_value(line)]
    for line_index, line in enumerate(lines):
        candidates = learning_field_candidates(field, line)
        for candidate_index, candidate in enumerate(candidates):
            if comparable_learning_value(field, candidate) == expected:
                return line_index, candidate_index
        if field not in {"amount", "eta", "invoice_date"}:
            if expected in comparable_learning_value(field, line):
                return line_index, None
    return None


def matching_template_line(old_lines: list[str], new_lines: list[str], old_index: int) -> str:
    if not new_lines:
        return ""
    expected_index = round(old_index * (len(new_lines) - 1) / max(len(old_lines) - 1, 1))
    old_signature = template_line_signature(old_lines[old_index])
    best_line = new_lines[min(expected_index, len(new_lines) - 1)]
    best_score = -1.0
    for index, line in enumerate(new_lines):
        line_score = SequenceMatcher(
            None, old_signature, template_line_signature(line), autojunk=False
        ).ratio()
        distance_penalty = abs(index - expected_index) / max(len(new_lines), 1) * 0.2
        score = line_score - distance_penalty
        if score > best_score:
            best_score = score
            best_line = line
    return best_line


def infer_field_from_example(field: str, corrected_value: str, old_text: str, new_text: str) -> str | None:
    if not corrected_value:
        return ""
    comparable = comparable_learning_value(field, corrected_value)
    if comparable and comparable in comparable_learning_value(field, new_text):
        return corrected_value
    if field in {"owner_name", "owner_email"}:
        return corrected_value

    location = locate_learning_value(field, corrected_value, old_text)
    if location is None:
        return None
    old_index, candidate_index = location
    if candidate_index is None:
        return None
    old_lines = [clean_value(line) for line in old_text.splitlines() if clean_value(line)]
    new_lines = [clean_value(line) for line in new_text.splitlines() if clean_value(line)]
    new_line = matching_template_line(old_lines, new_lines, old_index)
    candidates = learning_field_candidates(field, new_line)
    if not candidates:
        return None
    candidate = candidates[candidate_index] if candidate_index < len(candidates) else candidates[0]
    if field in {"eta", "invoice_date"}:
        return normalize_date(candidate)
    if field == "container_no":
        return re.sub(r"\s+", "", candidate).upper()
    if field == "currency":
        return candidate.upper()
    return candidate


def apply_learned_corrections(
    db: sqlite3.Connection, text: str, extracted_fields: dict[str, str]
) -> tuple[dict[str, str], str]:
    if not text.strip():
        return extracted_fields, ""
    signature = document_template_signature(text)
    examples = db.execute(
        "SELECT * FROM invoice_learning_examples ORDER BY updated_at DESC LIMIT 80"
    ).fetchall()
    if not examples:
        return extracted_fields, ""
    scored_examples = [
        (template_similarity(signature, example["template_signature"]), example)
        for example in examples
    ]
    score, example = max(scored_examples, key=lambda item: item[0])
    if score < 0.58:
        return extracted_fields, ""

    corrected = json.loads(example["fields_json"])
    confirmed_fields = json.loads(example["confirmed_fields"])
    learned_fields = []
    for field in confirmed_fields:
        learned_value = infer_field_from_example(
            field, corrected.get(field, ""), example["raw_text"], text
        )
        if learned_value is None:
            continue
        if extracted_fields.get(field, "") != learned_value:
            extracted_fields[field] = learned_value
            learned_fields.append(FIELD_LABELS.get(field, field))
    if not learned_fields:
        return extracted_fields, ""
    labels = "、".join(learned_fields)
    return extracted_fields, (
        f"已参考相似账单“{example['source_filename']}”的人工校正自动填写：{labels}"
        f"（版式相似度 {score:.0%}）。请继续核对。"
    )


def remember_invoice_learning(
    db: sqlite3.Connection, invoice: sqlite3.Row, corrected: dict[str, str]
) -> int:
    raw_text = invoice["raw_text"] or ""
    if not raw_text.strip():
        return 0
    confirmed_fields = []
    for field in TEMPLATE_LEARN_FIELDS:
        value = corrected.get(field, "")
        previous_value = invoice[field] or ""
        if value:
            if field in {"owner_name", "owner_email"} or locate_learning_value(field, value, raw_text):
                confirmed_fields.append(field)
        elif previous_value:
            confirmed_fields.append(field)
    if not confirmed_fields:
        return 0

    now = datetime.now().isoformat(timespec="seconds")
    values_to_store = {field: corrected.get(field, "") for field in TEMPLATE_LEARN_FIELDS}
    db.execute(
        """
        INSERT INTO invoice_learning_examples (
            invoice_id, source_filename, raw_text, template_signature,
            fields_json, confirmed_fields, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(invoice_id) DO UPDATE SET
            source_filename = excluded.source_filename,
            raw_text = excluded.raw_text,
            template_signature = excluded.template_signature,
            fields_json = excluded.fields_json,
            confirmed_fields = excluded.confirmed_fields,
            updated_at = excluded.updated_at
        """,
        (
            invoice["id"],
            invoice["source_filename"],
            raw_text,
            document_template_signature(raw_text),
            json.dumps(values_to_store, ensure_ascii=False),
            json.dumps(confirmed_fields, ensure_ascii=False),
            now,
            now,
        ),
    )
    return len(confirmed_fields)


def seed_learning_examples(db: sqlite3.Connection) -> None:
    invoices = db.execute(
        """
        SELECT i.* FROM invoices i
        LEFT JOIN invoice_learning_examples e ON e.invoice_id = i.id
        WHERE e.invoice_id IS NULL AND i.raw_text <> '' AND i.status <> '待校正'
        """
    ).fetchall()
    for invoice in invoices:
        remember_invoice_learning(
            db,
            invoice,
            {field: invoice[field] or "" for field in TEMPLATE_LEARN_FIELDS},
        )


def get_invoice(invoice_id: int) -> sqlite3.Row:
    with db_session() as db:
        invoice = db.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
    if invoice is None:
        abort(404)
    return invoice


def group_filters() -> tuple[str, list[str]]:
    clauses, values = [], []
    for column, argument in (("vessel_name", "vessel"),):
        value = request.args.get(argument)
        if value is not None:
            clauses.append(f"{column} = ?")
            values.append(value)
    where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
    return where, values


def split_container_numbers(value: str) -> list[str]:
    if not value:
        return []
    container_numbers = re.findall(r"\b[A-Z]{4}\d{7}\b", value.upper())
    if not container_numbers:
        container_numbers = [
            part.strip()
            for part in re.split(r"[,;/\n]+", value)
            if part.strip()
        ]
    return list(dict.fromkeys(container_numbers))


def is_cosco_owner(owner_name: str) -> bool:
    return "COSCO" in (owner_name or "").upper()


def load_vessel_groups(db: sqlite3.Connection, archived: bool = False) -> list[dict]:
    archive_join = (
        "JOIN vessel_archives a ON a.vessel_name = i.vessel_name"
        if archived
        else "LEFT JOIN vessel_archives a ON a.vessel_name = i.vessel_name"
    )
    archive_filter = "a.vessel_name IS NOT NULL" if archived else "a.vessel_name IS NULL"
    group_rows = db.execute(
        f"""
        SELECT i.vessel_name, MAX(NULLIF(i.eta, '')) AS eta, COUNT(*) AS invoice_count,
               COUNT(DISTINCT i.owner_name || char(31) || i.owner_email) AS owner_count,
               COUNT(DISTINCT CASE WHEN UPPER(i.owner_name) LIKE '%COSCO%'
                     THEN i.owner_name || char(31) || i.owner_email END) AS request_owner_count,
               SUM(CASE WHEN i.status = '已完成' THEN 1 ELSE 0 END) AS completed_count,
               MAX(i.updated_at) AS updated_at,
               MAX(a.archived_at) AS archived_at
        FROM invoices i
        {archive_join}
        WHERE {archive_filter}
        GROUP BY i.vessel_name
        ORDER BY CASE WHEN MAX(NULLIF(i.eta, '')) IS NULL THEN 1 ELSE 0 END,
                 MAX(NULLIF(i.eta, '')) ASC,
                 i.vessel_name ASC
        """
    ).fetchall()
    draft_counts = {
        row["vessel_name"]: row["draft_count"]
        for row in db.execute(
            """SELECT vessel_name, COUNT(*) AS draft_count
               FROM email_drafts
               WHERE UPPER(owner_name) LIKE '%COSCO%'
               GROUP BY vessel_name"""
        ).fetchall()
    }
    container_rows = db.execute(
        f"""
        SELECT i.vessel_name, i.owner_name, i.container_no
        FROM invoices i
        {archive_join}
        WHERE {archive_filter} AND i.container_no <> ''
        ORDER BY i.vessel_name, i.owner_name, i.container_no
        """
    ).fetchall()
    containers_by_vessel: dict[str, dict[str, list[str]]] = {}
    for container_row in container_rows:
        vessel_containers = containers_by_vessel.setdefault(container_row["vessel_name"], {})
        owner_containers = vessel_containers.setdefault(container_row["owner_name"], [])
        for container_number in split_container_numbers(container_row["container_no"]):
            if container_number not in owner_containers:
                owner_containers.append(container_number)
    today = date.today()
    groups = []
    for row in group_rows:
        group = dict(row)
        group["draft_count"] = draft_counts.get(row["vessel_name"], 0)
        group["container_groups"] = [
            {"owner_name": owner_name or "待补充船东", "containers": containers}
            for owner_name, containers in containers_by_vessel.get(row["vessel_name"], {}).items()
        ]
        group["days_to_eta"] = None
        if row["eta"]:
            try:
                group["days_to_eta"] = (datetime.strptime(row["eta"], "%Y-%m-%d").date() - today).days
            except ValueError:
                pass
        groups.append(group)
    return groups


@app.route("/")
def dashboard():
    with db_session() as db:
        groups = load_vessel_groups(db, archived=False)
    return render_template("dashboard.html", groups=groups)


@app.route("/history")
def history():
    with db_session() as db:
        groups = load_vessel_groups(db, archived=True)
    return render_template("history.html", groups=groups)


@app.route("/group/archive", methods=["POST"])
def archive_group():
    vessel = request.form.get("vessel", "").strip()
    if not vessel:
        abort(400)
    with db_session() as db:
        exists = db.execute("SELECT 1 FROM invoices WHERE vessel_name = ?", (vessel,)).fetchone()
        if not exists:
            abort(404)
        db.execute(
            """
            INSERT INTO vessel_archives (vessel_name, archived_at) VALUES (?, ?)
            ON CONFLICT(vessel_name) DO UPDATE SET archived_at = excluded.archived_at
            """,
            (vessel, datetime.now().isoformat(timespec="seconds")),
        )
    flash(f"{vessel} 已结束并移入历史记录。", "success")
    return redirect(url_for("dashboard"))


@app.route("/group/restore", methods=["POST"])
def restore_group():
    vessel = request.form.get("vessel", "").strip()
    if not vessel:
        abort(400)
    with db_session() as db:
        db.execute("DELETE FROM vessel_archives WHERE vessel_name = ?", (vessel,))
    flash(f"{vessel} 已恢复到当前船舶。", "success")
    return redirect(url_for("dashboard"))


@app.route("/upload", methods=["GET", "POST"])
def upload():
    if request.method == "GET":
        return render_template("upload.html")
    files = request.files.getlist("pdfs")
    valid_files = [file for file in files if file and file.filename]
    if not valid_files:
        flash("请选择至少一个 PDF 文件。", "error")
        return redirect(url_for("upload"))
    created_ids = []
    duplicate_ids = []
    for file in valid_files:
        source_name = file.filename.replace("\\", "/").rsplit("/", 1)[-1].replace("\x00", "").strip()
        if Path(source_name).suffix.lower() != ".pdf":
            flash(f"已跳过非 PDF 文件：{file.filename}", "error")
            continue
        stored_name = f"{uuid.uuid4().hex}.pdf"
        destination = UPLOAD_DIR / stored_name
        file.save(destination)
        text, warning = extract_pdf_text(destination)
        fields = extract_fields(text)
        now = datetime.now().isoformat(timespec="seconds")
        duplicate = None
        with db_session() as db:
            fields, learning_note = apply_learned_corrections(db, text, fields)
            duplicate = find_duplicate_invoice(db, fields["invoice_no"])
            if duplicate is None:
                cursor = db.execute(
                    """
                    INSERT INTO invoices (
                        source_filename, stored_filename, raw_text, extraction_warning, learning_note,
                        vessel_name, voyage_no, eta, invoice_no, invoice_date, owner_name, owner_email,
                        port_of_discharge, container_no, container_size, bl_no, charge_details,
                        amount, currency, status, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '待校正', ?, ?)
                    """,
                    (
                        source_name or "invoice.pdf",
                        stored_name,
                        text,
                        warning,
                        learning_note,
                        fields["vessel_name"],
                        fields["voyage_no"],
                        fields["eta"],
                        fields["invoice_no"],
                        fields["invoice_date"],
                        fields["owner_name"],
                        fields["owner_email"],
                        fields["port_of_discharge"],
                        fields["container_no"],
                        fields["container_size"],
                        fields["bl_no"],
                        fields["charge_details"],
                        fields["amount"],
                        fields["currency"],
                        now,
                        now,
                    ),
                )
                created_ids.append(cursor.lastrowid)
        if duplicate is not None:
            destination.unlink(missing_ok=True)
            duplicate_ids.append(duplicate["id"])
            flash(
                f"账单号 {fields['invoice_no']} 已经存在，已跳过重复文件并保留原记录。",
                "success",
            )
    if not created_ids:
        if duplicate_ids:
            return redirect(url_for("edit_invoice", invoice_id=duplicate_ids[0]))
        return redirect(url_for("upload"))
    flash(f"已导入 {len(created_ids)} 份账单，请核对识别结果。", "success")
    return redirect(url_for("edit_invoice", invoice_id=created_ids[0]))


@app.route("/invoice/<int:invoice_id>", methods=["GET", "POST"])
def edit_invoice(invoice_id: int):
    invoice = get_invoice(invoice_id)
    if request.method == "POST":
        corrected = {
            field: clean_value(request.form.get(field, "")) for field in EDITABLE_FIELDS
        }
        if is_cosco_owner(corrected["owner_name"]) and not corrected["owner_email"]:
            corrected["owner_email"] = COSCO_RELEASES_EMAIL
        with db_session() as db:
            duplicate = find_duplicate_invoice(db, corrected["invoice_no"], exclude_id=invoice_id)
            if duplicate is not None and invoice["status"] == "待校正":
                db.execute("DELETE FROM invoice_learning_examples WHERE invoice_id = ?", (invoice_id,))
                db.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))
            elif duplicate is not None:
                flash(
                    f"账单号 {corrected['invoice_no']} 已经存在，本次修改未保存。",
                    "error",
                )
                return redirect(url_for("edit_invoice", invoice_id=invoice_id))
        if duplicate is not None:
            duplicate_file = UPLOAD_DIR / invoice["stored_filename"]
            duplicate_file.unlink(missing_ok=True)
            flash(
                f"账单号 {corrected['invoice_no']} 已经存在，已丢弃刚导入的重复记录。",
                "success",
            )
            return redirect(url_for("edit_invoice", invoice_id=duplicate["id"]))
        values = [corrected[field] for field in EDITABLE_FIELDS]
        values.extend([datetime.now().isoformat(timespec="seconds"), invoice_id])
        assignments = ", ".join(f"{field} = ?" for field in EDITABLE_FIELDS)
        with db_session() as db:
            learned_count = remember_invoice_learning(db, invoice, corrected)
            db.execute(f"UPDATE invoices SET {assignments}, updated_at = ? WHERE id = ?", values)
        if learned_count:
            flash(f"账单信息已保存，系统已记住这类 PDF 的 {learned_count} 个字段位置。", "success")
        else:
            flash("账单信息已保存。", "success")
        if request.form.get("next") == "group":
            return redirect(url_for("invoice_group", vessel=request.form.get("vessel_name", "")))
        return redirect(url_for("edit_invoice", invoice_id=invoice_id))
    return render_template("edit_invoice.html", invoice=invoice)


def invoice_totals(invoices: list[sqlite3.Row]) -> list[dict[str, str]]:
    totals: dict[str, Decimal] = {}
    for invoice in invoices:
        currency = (invoice["currency"] or "未注明币种").upper()
        try:
            value = Decimal((invoice["amount"] or "0").replace(",", ""))
        except InvalidOperation:
            continue
        totals[currency] = totals.get(currency, Decimal("0")) + value
    return [{"currency": currency, "amount": f"{amount:.2f}"} for currency, amount in totals.items()]


def email_date(value: str) -> str:
    if not value:
        return "[ETA to be confirmed]"
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return value


def money_number(value: str) -> str:
    try:
        return f"{Decimal((value or '0').replace(',', '')):,.2f}"
    except InvalidOperation:
        return value or "0.00"


def invoice_value_label(invoice: sqlite3.Row) -> str:
    symbol = {"GBP": "£", "USD": "$", "EUR": "€", "CNY": "¥", "RMB": "¥"}.get(
        (invoice["currency"] or "").upper(),
        f"{invoice['currency']} ",
    )
    return f"{symbol}{money_number(invoice['amount'])}"


def release_subject_total(totals: list[dict[str, str]]) -> str:
    return " + ".join(f"{money_number(total['amount'])}{total['currency']}" for total in totals)


def payment_total_label(totals: list[dict[str, str]]) -> str:
    labels = []
    for total in totals:
        currency = total["currency"].upper()
        symbol = {"GBP": "£", "USD": "$", "EUR": "€", "CNY": "¥", "RMB": "¥"}.get(currency, f"{currency} ")
        labels.append(f"{symbol}{money_number(total['amount'])}")
    return " + ".join(labels)


def calculate_extension_days(last_free_day: str, pickup_date: str) -> int:
    if not last_free_day or not pickup_date:
        return 0
    try:
        free_date = datetime.strptime(last_free_day, "%Y-%m-%d").date()
        collection_date = datetime.strptime(pickup_date, "%Y-%m-%d").date()
    except ValueError:
        return 0
    return max((collection_date - free_date).days, 0)


def build_release_rows(invoices: list[sqlite3.Row], payment_date: str) -> list[dict[str, str]]:
    rows = []
    for invoice in invoices:
        containers = split_container_numbers(invoice["container_no"]) or [invoice["container_no"] or "-"]
        for container in containers:
            rows.append(
                {
                    "payment_date": email_date(payment_date),
                    "invoice_value": invoice_value_label(invoice),
                    "invoice_no": invoice["invoice_no"] or "-",
                    "invoice_date": email_date(invoice["invoice_date"]),
                    "vessel": invoice["vessel_name"],
                    "eta": email_date(invoice["eta"]),
                    "bl_no": invoice["bl_no"] or "-",
                    "container_no": container,
                }
            )
    return rows


def owner_summary_groups(
    invoices: list[sqlite3.Row],
    drafts: dict[tuple[str, str], sqlite3.Row],
    payment_proofs: dict[str, list[sqlite3.Row]] | None = None,
    release_drafts: dict[tuple[str, str], sqlite3.Row] | None = None,
    container_releases: dict[tuple[str, str], bool] | None = None,
    container_free_days: dict[tuple[str, str], sqlite3.Row] | None = None,
    extension_drafts: dict[tuple[str, str], sqlite3.Row] | None = None,
) -> list[dict]:
    payment_proofs = payment_proofs or {}
    release_drafts = release_drafts or {}
    container_releases = container_releases or {}
    container_free_days = container_free_days or {}
    extension_drafts = extension_drafts or {}
    groups: dict[tuple[str, str], dict] = {}
    for invoice in invoices:
        key = (invoice["owner_name"], invoice["owner_email"])
        if key not in groups:
            groups[key] = {
                "owner_name": invoice["owner_name"],
                "owner_email": invoice["owner_email"],
                "invoices": [],
                "containers": [],
                "container_map": {},
                "is_cosco": is_cosco_owner(invoice["owner_name"]),
                "draft": drafts.get(key),
                "release_draft": release_drafts.get(key),
                "extension_draft": extension_drafts.get(key),
                "payment_proofs": payment_proofs.get(invoice["owner_name"], []),
            }
        groups[key]["invoices"].append(invoice)
        container_numbers = split_container_numbers(invoice["container_no"]) or ["待补充箱号"]
        for container_number in container_numbers:
            if container_number not in groups[key]["containers"]:
                groups[key]["containers"].append(container_number)
            container_summary = groups[key]["container_map"].setdefault(
                container_number,
                {
                    "container_no": container_number,
                    "invoices": [],
                    "bl_nos": [],
                    "sizes": [],
                    "released": bool(container_releases.get((invoice["owner_name"], container_number), False)),
                },
            )
            container_summary["invoices"].append(invoice)
            if invoice["bl_no"] and invoice["bl_no"] not in container_summary["bl_nos"]:
                container_summary["bl_nos"].append(invoice["bl_no"])
            if invoice["container_size"] and invoice["container_size"] not in container_summary["sizes"]:
                container_summary["sizes"].append(invoice["container_size"])
    for group in groups.values():
        group["totals"] = invoice_totals(group["invoices"])
        group["container_summaries"] = []
        for container_summary in group.pop("container_map").values():
            free_day_plan = container_free_days.get(
                (group["owner_name"], container_summary["container_no"])
            )
            container_summary["last_free_day"] = (
                free_day_plan["last_free_day"] if free_day_plan else ""
            )
            container_summary["pickup_date"] = (
                free_day_plan["pickup_date"] if free_day_plan else ""
            )
            container_summary["extension_days"] = calculate_extension_days(
                container_summary["last_free_day"], container_summary["pickup_date"]
            )
            container_summary["needs_extension"] = container_summary["extension_days"] > 0
            container_summary["invoice_numbers"] = list(
                dict.fromkeys(
                    invoice["invoice_no"] or "未填写账单号"
                    for invoice in container_summary["invoices"]
                )
            )
            container_summary["totals"] = invoice_totals(container_summary["invoices"])
            group["container_summaries"].append(container_summary)
        group["released_count"] = sum(
            1 for container in group["container_summaries"] if container["released"]
        )
        group["extension_count"] = sum(
            1 for container in group["container_summaries"] if container["needs_extension"]
        )
        group["extension_invoice_count"] = sum(
            1
            for invoice in group["invoices"]
            if invoice["invoice_category"] == "last_free_day_extension"
        )
        last_free_days = {
            container["last_free_day"]
            for container in group["container_summaries"]
            if container["last_free_day"]
        }
        pickup_dates = {
            container["pickup_date"]
            for container in group["container_summaries"]
            if container["pickup_date"]
        }
        group["bulk_last_free_day"] = (
            next(iter(last_free_days))
            if len(last_free_days) == 1
            and len(last_free_days) == len({c["last_free_day"] for c in group["container_summaries"]})
            else ""
        )
        group["bulk_pickup_date"] = (
            next(iter(pickup_dates))
            if len(pickup_dates) == 1
            and len(pickup_dates) == len({c["pickup_date"] for c in group["container_summaries"]})
            else ""
        )
    return list(groups.values())


@app.route("/group")
def invoice_group():
    vessel = request.args.get("vessel", "").strip()
    if not vessel:
        return redirect(url_for("dashboard"))
    with db_session() as db:
        invoices = db.execute(
            "SELECT * FROM invoices WHERE vessel_name = ? ORDER BY owner_name, owner_email, invoice_no",
            (vessel,),
        ).fetchall()
        draft_rows = db.execute(
            "SELECT * FROM email_drafts WHERE vessel_name = ?",
            (vessel,),
        ).fetchall()
        proof_rows = db.execute(
            "SELECT * FROM payment_proofs WHERE vessel_name = ? ORDER BY created_at DESC, id DESC",
            (vessel,),
        ).fetchall()
        release_rows = db.execute(
            "SELECT * FROM cosco_release_drafts WHERE vessel_name = ?",
            (vessel,),
        ).fetchall()
        container_release_rows = db.execute(
            "SELECT * FROM container_releases WHERE vessel_name = ?",
            (vessel,),
        ).fetchall()
        free_day_rows = db.execute(
            "SELECT * FROM container_free_days WHERE vessel_name = ?",
            (vessel,),
        ).fetchall()
        extension_draft_rows = db.execute(
            "SELECT * FROM free_day_extension_drafts WHERE vessel_name = ?",
            (vessel,),
        ).fetchall()
        archived = db.execute(
            "SELECT 1 FROM vessel_archives WHERE vessel_name = ?",
            (vessel,),
        ).fetchone() is not None
    drafts = {(row["owner_name"], row["owner_email"]): row for row in draft_rows}
    proofs_by_owner: dict[str, list[sqlite3.Row]] = {}
    for proof in proof_rows:
        proofs_by_owner.setdefault(proof["owner_name"], []).append(proof)
    releases = {(row["owner_name"], row["owner_email"]): row for row in release_rows}
    container_releases = {
        (row["owner_name"], row["container_no"]): bool(row["released"])
        for row in container_release_rows
    }
    free_day_plans = {
        (row["owner_name"], row["container_no"]): row for row in free_day_rows
    }
    extension_drafts = {
        (row["owner_name"], row["owner_email"]): row
        for row in extension_draft_rows
    }
    owner_groups = owner_summary_groups(
        invoices,
        drafts,
        proofs_by_owner,
        releases,
        container_releases,
        free_day_plans,
        extension_drafts,
    )
    eta = next((invoice["eta"] for invoice in invoices if invoice["eta"]), "")
    return render_template(
        "group.html",
        vessel=vessel,
        eta=eta,
        invoices=invoices,
        owner_groups=owner_groups,
        grand_totals=invoice_totals(invoices),
        archived=archived,
        today_iso=date.today().isoformat(),
    )


@app.route("/group/eta", methods=["POST"])
def update_group_eta():
    vessel = request.form.get("vessel", "").strip()
    eta = request.form.get("eta", "").strip()
    if not vessel:
        abort(400)
    with db_session() as db:
        db.execute(
            "UPDATE invoices SET eta = ?, updated_at = ? WHERE vessel_name = ?",
            (eta, datetime.now().isoformat(timespec="seconds"), vessel),
        )
    flash("这条船的 ETA 已统一更新。", "success")
    return redirect(url_for("invoice_group", vessel=vessel))


@app.route("/group/free-days", methods=["POST"])
def update_container_free_days():
    vessel = request.form.get("vessel", "").strip()
    owner = request.form.get("owner", "").strip()
    scope = request.form.get("scope", "container").strip()
    container_no = request.form.get("container", "").strip()
    last_free_day = request.form.get("last_free_day", "").strip()
    pickup_date = request.form.get("pickup_date", "").strip()
    if not vessel or not owner or scope not in {"owner", "container"}:
        abort(400)
    for value in (last_free_day, pickup_date):
        if value:
            try:
                datetime.strptime(value, "%Y-%m-%d")
            except ValueError:
                flash("日期格式无效，请重新选择。", "error")
                return redirect(url_for("invoice_group", vessel=vessel))

    with db_session() as db:
        invoice_rows = db.execute(
            "SELECT container_no FROM invoices WHERE vessel_name = ? AND owner_name = ?",
            (vessel, owner),
        ).fetchall()
        valid_containers = list(
            dict.fromkeys(
                number
                for invoice in invoice_rows
                for number in split_container_numbers(invoice["container_no"])
            )
        )
        if not valid_containers:
            abort(404)
        targets = valid_containers if scope == "owner" else [container_no]
        if any(target not in valid_containers for target in targets):
            abort(404)
        now = datetime.now().isoformat(timespec="seconds")
        for target in targets:
            db.execute(
                """
                INSERT INTO container_free_days (
                    vessel_name, owner_name, container_no, last_free_day, pickup_date, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(vessel_name, owner_name, container_no) DO UPDATE SET
                    last_free_day = excluded.last_free_day,
                    pickup_date = excluded.pickup_date,
                    updated_at = excluded.updated_at
                """,
                (vessel, owner, target, last_free_day, pickup_date, now),
            )
    if scope == "owner":
        flash(f"已为 {owner} 的 {len(targets)} 个集装箱统一更新免租期计划。", "success")
    else:
        flash(f"已更新集装箱 {container_no} 的免租期计划。", "success")
    return redirect(url_for("invoice_group", vessel=vessel))


@app.route("/group/free-day-draft", methods=["GET", "POST"])
def free_day_extension_draft():
    vessel = request.args.get("vessel", "").strip()
    owner = request.args.get("owner", "").strip()
    email = request.args.get("email", "").strip()
    if not vessel or not owner:
        abort(400)
    with db_session() as db:
        invoice_rows = db.execute(
            "SELECT container_no FROM invoices WHERE vessel_name = ? AND owner_name = ?",
            (vessel, owner),
        ).fetchall()
        valid_containers = list(
            dict.fromkeys(
                number
                for invoice in invoice_rows
                for number in split_container_numbers(invoice["container_no"])
            )
        )
        plans = db.execute(
            """
            SELECT * FROM container_free_days
            WHERE vessel_name = ? AND owner_name = ?
            ORDER BY container_no
            """,
            (vessel, owner),
        ).fetchall()
        saved_draft = db.execute(
            """
            SELECT * FROM free_day_extension_drafts
            WHERE vessel_name = ? AND owner_name = ? AND owner_email = ?
            """,
            (vessel, owner, email),
        ).fetchone()
    if not valid_containers:
        abort(404)
    extension_rows = [
        {
            "container_no": plan["container_no"],
            "last_free_day": plan["last_free_day"],
            "pickup_date": plan["pickup_date"],
            "pickup_date_display": email_date(plan["pickup_date"]),
            "extension_days": calculate_extension_days(
                plan["last_free_day"], plan["pickup_date"]
            ),
        }
        for plan in plans
        if plan["container_no"] in valid_containers
        and calculate_extension_days(plan["last_free_day"], plan["pickup_date"]) > 0
    ]
    if not extension_rows:
        flash("目前没有需要延长 Last Free Day 的集装箱。", "error")
        return redirect(url_for("invoice_group", vessel=vessel))

    default_subject = f"Last Free Day Extension - {vessel} // {DEFAULT_COMPANY_NAME}"
    default_body = render_template(
        "free_day_extension_email.txt", containers=extension_rows
    )
    formatted_html = render_template(
        "free_day_extension_email.html", containers=extension_rows
    )
    if request.method == "POST":
        subject = request.form.get("email_subject", "").strip()
        body = request.form.get("email_body", "").strip()
        now = datetime.now().isoformat(timespec="seconds")
        with db_session() as db:
            db.execute(
                """
                INSERT INTO free_day_extension_drafts (
                    vessel_name, owner_name, owner_email, subject, body, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(vessel_name, owner_name, owner_email) DO UPDATE SET
                    subject = excluded.subject,
                    body = excluded.body,
                    updated_at = excluded.updated_at
                """,
                (vessel, owner, email, subject, body, now, now),
            )
        flash("Last Free Day 延期邮件草稿已保存。", "success")
        return redirect(
            url_for(
                "free_day_extension_draft", vessel=vessel, owner=owner, email=email
            )
        )
    return render_template(
        "free_day_extension_draft.html",
        vessel=vessel,
        owner=owner,
        email=email,
        containers=extension_rows,
        formatted_html=formatted_html,
        subject=saved_draft["subject"] if saved_draft else default_subject,
        body=saved_draft["body"] if saved_draft else default_body,
    )


@app.route("/group/extension-invoice", methods=["POST"])
def upload_extension_invoice():
    vessel = request.form.get("vessel", "").strip()
    owner = request.form.get("owner", "").strip()
    email = request.form.get("email", "").strip()
    container_hint = request.form.get("container_hint", "").strip()
    files = [file for file in request.files.getlist("extension_pdfs") if file and file.filename]
    if not vessel or not owner:
        abort(400)
    if not files:
        flash("请选择至少一份延长免租期账单 PDF。", "error")
        return redirect(url_for("invoice_group", vessel=vessel))

    with db_session() as db:
        related_invoices = db.execute(
            "SELECT * FROM invoices WHERE vessel_name = ? AND owner_name = ? ORDER BY id",
            (vessel, owner),
        ).fetchall()
    if not related_invoices:
        abort(404)
    valid_containers = list(
        dict.fromkeys(
            number
            for invoice in related_invoices
            for number in split_container_numbers(invoice["container_no"])
        )
    )
    if container_hint and container_hint not in valid_containers:
        abort(400)
    eta = next((invoice["eta"] for invoice in related_invoices if invoice["eta"]), "")
    created_ids, duplicate_ids = [], []
    for file in files:
        source_name = file.filename.replace("\\", "/").rsplit("/", 1)[-1].replace("\x00", "").strip()
        if Path(source_name).suffix.lower() != ".pdf":
            flash(f"已跳过非 PDF 文件：{source_name}", "error")
            continue
        stored_name = f"{uuid.uuid4().hex}.pdf"
        destination = UPLOAD_DIR / stored_name
        file.save(destination)
        text, warning = extract_pdf_text(destination)
        fields = extract_fields(text)
        now = datetime.now().isoformat(timespec="seconds")
        duplicate = None
        with db_session() as db:
            fields, learning_note = apply_learned_corrections(db, text, fields)
            fields["vessel_name"] = vessel
            fields["owner_name"] = owner
            fields["owner_email"] = email
            fields["eta"] = fields["eta"] or eta
            if container_hint:
                fields["container_no"] = container_hint
            duplicate = find_duplicate_invoice(db, fields["invoice_no"])
            if duplicate is None:
                cursor = db.execute(
                    """
                    INSERT INTO invoices (
                        source_filename, stored_filename, raw_text, extraction_warning,
                        learning_note, invoice_category, vessel_name, voyage_no, eta,
                        invoice_no, invoice_date, owner_name, owner_email, port_of_discharge,
                        container_no, container_size, bl_no, charge_details, amount, currency,
                        status, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, 'last_free_day_extension', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '待校正', ?, ?)
                    """,
                    (
                        source_name or "extension-invoice.pdf",
                        stored_name,
                        text,
                        warning,
                        learning_note,
                        fields["vessel_name"], fields["voyage_no"], fields["eta"],
                        fields["invoice_no"], fields["invoice_date"], fields["owner_name"],
                        fields["owner_email"], fields["port_of_discharge"], fields["container_no"],
                        fields["container_size"], fields["bl_no"], fields["charge_details"],
                        fields["amount"], fields["currency"], now, now,
                    ),
                )
                created_ids.append(cursor.lastrowid)
        if duplicate is not None:
            destination.unlink(missing_ok=True)
            duplicate_ids.append(duplicate["id"])
            flash(
                f"账单号 {fields['invoice_no']} 已存在，已跳过重复的延期账单。",
                "success",
            )
    if created_ids:
        flash(f"已导入 {len(created_ids)} 份延长免租期账单，请核对后再付款。", "success")
        return redirect(url_for("edit_invoice", invoice_id=created_ids[0]))
    if duplicate_ids:
        return redirect(url_for("edit_invoice", invoice_id=duplicate_ids[0]))
    return redirect(url_for("invoice_group", vessel=vessel))


@app.route("/container/release", methods=["POST"])
def update_container_release():
    vessel = request.form.get("vessel", "").strip()
    owner = request.form.get("owner", "").strip()
    container_no = request.form.get("container", "").strip()
    released = request.form.get("released", "0") == "1"
    if not vessel or not owner or not container_no:
        abort(400)

    with db_session() as db:
        invoices = db.execute(
            "SELECT container_no FROM invoices WHERE vessel_name = ? AND owner_name = ?",
            (vessel, owner),
        ).fetchall()
        valid_containers = {
            number
            for invoice in invoices
            for number in split_container_numbers(invoice["container_no"])
        }
        if container_no not in valid_containers:
            abort(404)
        now = datetime.now().isoformat(timespec="seconds")
        db.execute(
            """
            INSERT INTO container_releases (
                vessel_name, owner_name, container_no, released, updated_at
            ) VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(vessel_name, owner_name, container_no) DO UPDATE SET
                released = excluded.released,
                updated_at = excluded.updated_at
            """,
            (vessel, owner, container_no, int(released), now),
        )
    return jsonify(ok=True, released=released)


@app.route("/group/payment-proof", methods=["POST"])
def upload_payment_proof():
    vessel = request.form.get("vessel", "").strip()
    owner = request.form.get("owner", "").strip()
    payment_date = request.form.get("payment_date", "").strip()
    files = [file for file in request.files.getlist("proof_files") if file and file.filename]
    if not vessel or not owner:
        abort(400)
    if not files:
        flash("请选择至少一张付款截图。", "error")
        return redirect(url_for("invoice_group", vessel=vessel))
    with db_session() as db:
        exists = db.execute(
            "SELECT 1 FROM invoices WHERE vessel_name = ? AND owner_name = ?",
            (vessel, owner),
        ).fetchone()
        if not exists:
            abort(404)
        saved_count = 0
        now = datetime.now().isoformat(timespec="seconds")
        for file in files:
            source_name = file.filename.replace("\\", "/").rsplit("/", 1)[-1].replace("\x00", "").strip()
            suffix = Path(source_name).suffix.lower()
            if suffix not in PAYMENT_PROOF_SUFFIXES:
                flash(f"已跳过不支持的截图格式：{source_name}", "error")
                continue
            stored_name = f"{uuid.uuid4().hex}{suffix}"
            file.save(PAYMENT_PROOF_DIR / stored_name)
            db.execute(
                """
                INSERT INTO payment_proofs (
                    vessel_name, owner_name, payment_date, source_filename, stored_filename, created_at
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (vessel, owner, payment_date, source_name, stored_name, now),
            )
            saved_count += 1
    if saved_count:
        flash(f"已为 {owner} 保存 {saved_count} 张付款截图。", "success")
    return redirect(url_for("invoice_group", vessel=vessel))


@app.route("/payment-proof/<int:proof_id>/file")
def payment_proof_file(proof_id: int):
    with db_session() as db:
        proof = db.execute("SELECT * FROM payment_proofs WHERE id = ?", (proof_id,)).fetchone()
    if proof is None:
        abort(404)
    path = PAYMENT_PROOF_DIR / proof["stored_filename"]
    if not path.is_file():
        abort(404)
    mimetype = mimetypes.guess_type(proof["source_filename"])[0] or "application/octet-stream"
    return send_file(path, mimetype=mimetype, download_name=proof["source_filename"], as_attachment=False)


@app.route("/payment-proof/<int:proof_id>/delete", methods=["POST"])
def delete_payment_proof(proof_id: int):
    with db_session() as db:
        proof = db.execute("SELECT * FROM payment_proofs WHERE id = ?", (proof_id,)).fetchone()
        if proof is None:
            abort(404)
        db.execute("DELETE FROM payment_proofs WHERE id = ?", (proof_id,))
    path = PAYMENT_PROOF_DIR / proof["stored_filename"]
    path.unlink(missing_ok=True)
    flash("付款截图已移除。", "success")
    return redirect(url_for("invoice_group", vessel=proof["vessel_name"]))


@app.route("/invoice/<int:invoice_id>/draft", methods=["GET", "POST"])
def email_draft(invoice_id: int):
    invoice = get_invoice(invoice_id)
    return redirect(
        url_for(
            "group_email_draft",
            vessel=invoice["vessel_name"],
            owner=invoice["owner_name"],
            email=invoice["owner_email"],
        )
    )


@app.route("/group/draft", methods=["GET", "POST"])
def group_email_draft():
    vessel = request.args.get("vessel", "").strip()
    owner = request.args.get("owner", "").strip()
    email = request.args.get("email", "").strip()
    if not vessel:
        abort(400)
    with db_session() as db:
        invoices = db.execute(
            """
            SELECT * FROM invoices
            WHERE vessel_name = ? AND owner_name = ? AND owner_email = ?
            ORDER BY invoice_no, container_no
            """,
            (vessel, owner, email),
        ).fetchall()
        saved_draft = db.execute(
            "SELECT * FROM email_drafts WHERE vessel_name = ? AND owner_name = ? AND owner_email = ?",
            (vessel, owner, email),
        ).fetchone()
    if not invoices:
        abort(404)
    eta = next((invoice["eta"] for invoice in invoices if invoice["eta"]), "")
    totals = invoice_totals(invoices)
    is_cosco = is_cosco_owner(owner)
    if not is_cosco:
        flash("只有 COSCO 需要生成索要账单邮件；其他船东仅做账单汇总。", "error")
        return redirect(url_for("invoice_group", vessel=vessel))
    container_summaries = owner_summary_groups(invoices, {})[0]["container_summaries"]
    formatted_html = ""
    if is_cosco:
        default_subject = f"Release - {vessel} // {DEFAULT_COMPANY_NAME}"
        default_body = render_template(
            "cosco_release_email.txt",
            vessel=vessel,
            eta=email_date(eta),
            containers=container_summaries,
        )
        formatted_html = render_template(
            "cosco_release_email.html",
            vessel=vessel,
            eta=email_date(eta),
            containers=container_summaries,
        )
    else:
        default_subject = f"Vessel Invoice Summary - {vessel}"
        default_body = render_template(
            "group_email_body.txt",
            vessel=vessel,
            owner=owner,
            eta=eta,
            invoices=invoices,
            totals=totals,
        )
    if request.method == "POST":
        subject = request.form.get("email_subject", "").strip()
        body = request.form.get("email_body", "").strip()
        now = datetime.now().isoformat(timespec="seconds")
        with db_session() as db:
            db.execute(
                """
                INSERT INTO email_drafts (
                    vessel_name, owner_name, owner_email, subject, body, status, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, '草稿已保存', ?, ?)
                ON CONFLICT(vessel_name, owner_name, owner_email) DO UPDATE SET
                    subject = excluded.subject,
                    body = excluded.body,
                    status = excluded.status,
                    updated_at = excluded.updated_at
                """,
                (vessel, owner, email, subject, body, now, now),
            )
        flash("该船东的汇总邮件草稿已保存（本版本不会自动发送）。", "success")
        return redirect(url_for("group_email_draft", vessel=vessel, owner=owner, email=email))
    return render_template(
        "email_draft.html",
        vessel=vessel,
        owner=owner,
        email=email,
        invoices=invoices,
        totals=totals,
        is_cosco=is_cosco,
        formatted_html=formatted_html,
        subject=saved_draft["subject"] if saved_draft else default_subject,
        body=saved_draft["body"] if saved_draft else default_body,
    )


@app.route("/group/release-draft", methods=["GET", "POST"])
def cosco_release_draft():
    vessel = request.args.get("vessel", "").strip()
    owner = request.args.get("owner", "").strip()
    email = request.args.get("email", "").strip()
    if not vessel or not is_cosco_owner(owner):
        abort(400)
    with db_session() as db:
        invoices = db.execute(
            """
            SELECT * FROM invoices
            WHERE vessel_name = ? AND owner_name = ? AND owner_email = ?
            ORDER BY invoice_no, container_no
            """,
            (vessel, owner, email),
        ).fetchall()
        proofs = db.execute(
            """
            SELECT * FROM payment_proofs
            WHERE vessel_name = ? AND owner_name = ?
            ORDER BY created_at DESC, id DESC
            """,
            (vessel, owner),
        ).fetchall()
        saved_draft = db.execute(
            """
            SELECT * FROM cosco_release_drafts
            WHERE vessel_name = ? AND owner_name = ? AND owner_email = ?
            """,
            (vessel, owner, email),
        ).fetchone()
    if not invoices:
        abort(404)
    payment_date = next(
        (proof["payment_date"] for proof in proofs if proof["payment_date"]),
        date.today().isoformat(),
    )
    totals = invoice_totals(invoices)
    total_label = release_subject_total(totals)
    total_payment_value = payment_total_label(totals)
    release_rows = build_release_rows(invoices, payment_date)
    default_subject = f"Re: Release - {vessel} // Remittance {total_label} {DEFAULT_COMPANY_NAME}"
    default_body = render_template(
        "cosco_payment_release_email.txt",
        badge=DEFAULT_BADGE,
        rows=release_rows,
        total_label=total_label,
        total_payment_value=total_payment_value,
    )
    formatted_html = render_template(
        "cosco_payment_release_email.html",
        badge=DEFAULT_BADGE,
        rows=release_rows,
        total_label=total_label,
        total_payment_value=total_payment_value,
    )
    if request.method == "POST":
        subject = request.form.get("email_subject", "").strip()
        body = request.form.get("email_body", "").strip()
        now = datetime.now().isoformat(timespec="seconds")
        with db_session() as db:
            db.execute(
                """
                INSERT INTO cosco_release_drafts (
                    vessel_name, owner_name, owner_email, subject, body, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(vessel_name, owner_name, owner_email) DO UPDATE SET
                    subject = excluded.subject,
                    body = excluded.body,
                    updated_at = excluded.updated_at
                """,
                (vessel, owner, email, subject, body, now, now),
            )
        flash("COSCO 付款后 Release 邮件草稿已保存。", "success")
        return redirect(url_for("cosco_release_draft", vessel=vessel, owner=owner, email=email))
    return render_template(
        "cosco_release_draft.html",
        vessel=vessel,
        owner=owner,
        email=email,
        invoices=invoices,
        proofs=proofs,
        formatted_html=formatted_html,
        subject=saved_draft["subject"] if saved_draft else default_subject,
        body=saved_draft["body"] if saved_draft else default_body,
    )


@app.route("/invoice/<int:invoice_id>/file")
def invoice_file(invoice_id: int):
    invoice = get_invoice(invoice_id)
    path = UPLOAD_DIR / invoice["stored_filename"]
    if not path.is_file():
        abort(404)
    return send_file(path, mimetype="application/pdf", download_name=invoice["source_filename"], as_attachment=False)


@app.route("/export.xlsx")
def export_excel():
    where, values = group_filters()
    with db_session() as db:
        invoices = db.execute(f"SELECT * FROM invoices{where} ORDER BY eta, vessel_name, voyage_no", values).fetchall()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Shipowner Invoices"
    columns = [
        ("船名", "vessel_name"), ("航次", "voyage_no"), ("ETA", "eta"),
        ("船东", "owner_name"), ("邮箱", "owner_email"), ("集装箱号", "container_no"),
        ("卸货港", "port_of_discharge"), ("箱型", "container_size"), ("费用明细", "charge_details"),
        ("提单号", "bl_no"), ("账单号", "invoice_no"), ("账单日期", "invoice_date"), ("金额", "amount"),
        ("币种", "currency"), ("账单类型", "invoice_category"), ("状态", "status"), ("源文件", "source_filename"),
    ]
    sheet.append([title for title, _ in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="245B78")
    for invoice in invoices:
        sheet.append([invoice[key] for _, key in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [24, 14, 13, 24, 30, 22, 18, 12, 45, 18, 18, 14, 14, 10, 24, 14, 28]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"shipowner-invoices-{datetime.now():%Y%m%d}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.errorhandler(413)
def too_large(_error):
    flash("文件过大。单次上传总大小不能超过 20 MB。", "error")
    return redirect(url_for("upload"))


init_db()

if __name__ == "__main__":
    app_port = int(sys.argv[1] if len(sys.argv) > 1 else os.environ.get("SHIPOWNER_PORT", "5000"))
    if os.environ.get("SHIPOWNER_OPEN_BROWSER") == "1":
        threading.Timer(
            1.2, lambda: webbrowser.open(f"http://127.0.0.1:{app_port}")
        ).start()
    app.run(
        host="127.0.0.1",
        port=app_port,
        debug=os.environ.get("FLASK_DEBUG") == "1",
        use_reloader=False,
    )
